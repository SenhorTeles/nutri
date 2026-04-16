import os
import sys
import csv
import json
import threading
import tkinter as tk
import subprocess
from tkinter import ttk, messagebox, filedialog, font as tkfont
from datetime import datetime
import time

ORACLE_USUARIO = "MIGRACAO"
ORACLE_SENHA = "fzabu69128XPKGY@!"
ORACLE_HOST = "201.157.211.96"
ORACLE_PORTA = 1521
ORACLE_SERVICE = "CS8NZK_190797_W_high.paas.oracle.com"

CLIENT_LIB_DIR = r"C:\Users\informatica.ti\Documents\appdiscooveryzynapse\cmdintanci\instantclient_21_19"

COLORS = {
    "bg_dark": "#f5f6fa",
    "bg_medium": "#ffffff",
    "bg_light": "#ebedf2",
    "bg_input": "#ffffff",
    "accent": "#e67e22",
    "accent_hover": "#d35400",
    "accent_green": "#27ae60",
    "accent_red": "#e74c3c",
    "accent_orange": "#f39c12",
    "text_primary": "#2c3e50",
    "text_secondary": "#636e72",
    "text_muted": "#b2bec3",
    "border": "#dcdde1",
    "row_even": "#ffffff",
    "row_odd": "#fafafa",
    "row_hover": "#fef3e2",
    "selection": "#fdebd0",
    "scrollbar_bg": "#ebedf2",
    "scrollbar_fg": "#c8c8d0",
    "success_bg": "#eafaf1",
    "error_bg": "#fdedec",
}

SQL_KEYWORDS = [
    "SELECT",
    "FROM",
    "WHERE",
    "AND",
    "OR",
    "NOT",
    "IN",
    "BETWEEN",
    "LIKE",
    "IS",
    "NULL",
    "ORDER",
    "BY",
    "GROUP",
    "HAVING",
    "JOIN",
    "LEFT",
    "RIGHT",
    "INNER",
    "OUTER",
    "FULL",
    "CROSS",
    "ON",
    "AS",
    "INSERT",
    "INTO",
    "VALUES",
    "UPDATE",
    "SET",
    "DELETE",
    "CREATE",
    "TABLE",
    "ALTER",
    "DROP",
    "INDEX",
    "VIEW",
    "DISTINCT",
    "TOP",
    "LIMIT",
    "OFFSET",
    "UNION",
    "ALL",
    "EXISTS",
    "CASE",
    "WHEN",
    "THEN",
    "ELSE",
    "END",
    "COUNT",
    "SUM",
    "AVG",
    "MIN",
    "MAX",
    "FETCH",
    "FIRST",
    "ROWS",
    "ONLY",
    "WITH",
    "ROWNUM",
    "SYSDATE",
    "NVL",
    "NVL2",
    "DECODE",
    "TO_DATE",
    "TO_CHAR",
    "TO_NUMBER",
    "SUBSTR",
    "TRIM",
    "UPPER",
    "LOWER",
    "REPLACE",
    "INSTR",
    "ROUND",
    "TRUNC",
    "COALESCE",
    "DUAL",
    "ASC",
    "DESC",
]

HISTORY_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), ".vieworacle_history.json"
)


def log(msg):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}")


oracle_available = False
try:
    import oracledb
except ImportError:
    log("[ERRO] oracledb nao encontrado. O executavel nao foi compilado corretamente.")

try:
    import openpyxl
except ImportError:
    log("[ERRO] openpyxl nao encontrado. O executavel nao foi compilado corretamente.")

try:
    if "oracledb" in sys.modules:
        oracledb.init_oracle_client(lib_dir=CLIENT_LIB_DIR)
        log("[OK] Oracle Client inicializado!")
        oracle_available = True
except Exception as e:
    log(f"[AVISO] Oracle Client init: {e}")
    # Fallback para o Thin mode, que não precisa de instalação de cliente Oracle!
    oracle_available = True


def conectar():
    dsn = oracledb.makedsn(ORACLE_HOST, ORACLE_PORTA, service_name=ORACLE_SERVICE)
    connection = oracledb.connect(user=ORACLE_USUARIO, password=ORACLE_SENHA, dsn=dsn)
    return connection


class ViewOracleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("View WintxZyNapse — SQL Client · Winthor Online")
        self.root.geometry("1400x850")
        self.root.minsize(900, 600)
        self.root.configure(bg=COLORS["bg_dark"])

        self.connection = None
        self.current_columns = []
        self.current_data = []
        self.query_history = []
        self.is_running = False
        self.sort_column = None
        self.sort_reverse = False
        self.filter_text = tk.StringVar()
        self.filter_text.trace_add("write", self._on_filter_change)

        self._load_history()

        try:
            self.root.iconname("WintxZyNapse")
        except:
            pass

        self._setup_styles()
        self._build_ui()

        self.root.bind("<F5>", lambda e: self._execute_query())
        self.root.bind("<Control-Return>", lambda e: self._execute_query())
        self.root.bind("<Control-e>", lambda e: self._export_excel())
        self.root.bind("<Control-E>", lambda e: self._export_excel())
        self.root.bind("<Control-l>", lambda e: self._clear_editor())
        self.root.bind("<Control-L>", lambda e: self._clear_editor())

        self._set_status("Pronto. Pressione F5 ou Ctrl+Enter para executar.", "info")

    def _setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use("clam")

        self.style.configure(
            "Custom.Treeview",
            background=COLORS["bg_medium"],
            foreground=COLORS["text_primary"],
            fieldbackground=COLORS["bg_medium"],
            borderwidth=0,
            rowheight=38,
            font=("Segoe UI", 10),
        )
        self.style.configure(
            "Custom.Treeview.Heading",
            background=COLORS["bg_light"],
            foreground=COLORS["text_primary"],
            borderwidth=1,
            relief="flat",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 10),
        )
        self.style.map(
            "Custom.Treeview.Heading",
            background=[("active", COLORS["accent"])],
            foreground=[("active", "#ffffff")],
        )
        self.style.map(
            "Custom.Treeview",
            background=[("selected", COLORS["selection"])],
            foreground=[("selected", COLORS["text_primary"])],
        )

        self.style.configure(
            "Custom.Vertical.TScrollbar",
            background=COLORS["scrollbar_fg"],
            troughcolor=COLORS["scrollbar_bg"],
            borderwidth=0,
            arrowsize=0,
            width=10,
        )
        self.style.configure(
            "Custom.Horizontal.TScrollbar",
            background=COLORS["scrollbar_fg"],
            troughcolor=COLORS["scrollbar_bg"],
            borderwidth=0,
            arrowsize=0,
            width=10,
        )

        self.style.configure(
            "Accent.TButton",
            background=COLORS["accent"],
            foreground="#ffffff",
            borderwidth=0,
            padding=(16, 8),
            font=("Segoe UI", 10, "bold"),
        )
        self.style.map(
            "Accent.TButton",
            background=[
                ("active", COLORS["accent_hover"]),
                ("disabled", COLORS["bg_light"]),
            ],
        )

        self.style.configure(
            "Green.TButton",
            background=COLORS["accent_green"],
            foreground="#ffffff",
            borderwidth=0,
            padding=(16, 8),
            font=("Segoe UI", 10, "bold"),
        )
        self.style.map(
            "Green.TButton",
            background=[("active", "#00d1a0"), ("disabled", COLORS["bg_light"])],
        )

        self.style.configure(
            "Red.TButton",
            background=COLORS["accent_red"],
            foreground="#ffffff",
            borderwidth=0,
            padding=(16, 8),
            font=("Segoe UI", 10, "bold"),
        )
        self.style.map(
            "Red.TButton",
            background=[("active", "#ff5252"), ("disabled", COLORS["bg_light"])],
        )

        self.style.configure(
            "Flat.TButton",
            background=COLORS["bg_light"],
            foreground=COLORS["text_primary"],
            borderwidth=0,
            padding=(12, 8),
            font=("Segoe UI", 9),
        )
        self.style.map(
            "Flat.TButton",
            background=[("active", COLORS["border"])],
        )

        self.style.configure(
            "Custom.TEntry",
            fieldbackground=COLORS["bg_input"],
            foreground=COLORS["text_primary"],
            borderwidth=1,
            padding=(8, 6),
        )

        self.style.configure(
            "Custom.TNotebook",
            background=COLORS["bg_dark"],
            borderwidth=0,
        )
        self.style.configure(
            "Custom.TNotebook.Tab",
            background=COLORS["bg_light"],
            foreground=COLORS["text_secondary"],
            padding=(14, 6),
            font=("Segoe UI", 9),
        )
        self.style.map(
            "Custom.TNotebook.Tab",
            background=[("selected", COLORS["accent"])],
            foreground=[("selected", "#ffffff")],
        )

    def _build_ui(self):
        topbar = tk.Frame(self.root, bg=COLORS["bg_dark"], height=56)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)

        logo_frame = tk.Frame(topbar, bg=COLORS["bg_dark"])
        logo_frame.pack(side="left", padx=16)
        tk.Label(
            logo_frame,
            text="⬡",
            font=("Segoe UI", 22),
            fg=COLORS["accent"],
            bg=COLORS["bg_dark"],
        ).pack(side="left")
        tk.Label(
            logo_frame,
            text=" View",
            font=("Segoe UI", 16, "bold"),
            fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"],
        ).pack(side="left")
        tk.Label(
            logo_frame,
            text="Wintx",
            font=("Segoe UI", 16, "bold"),
            fg=COLORS["accent"],
            bg=COLORS["bg_dark"],
        ).pack(side="left")
        tk.Label(
            logo_frame,
            text="ZyNapse",
            font=("Segoe UI", 16, "bold"),
            fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"],
        ).pack(side="left")
        tk.Label(
            logo_frame,
            text="  SQL Client",
            font=("Segoe UI", 10),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"],
        ).pack(side="left", padx=(4, 0))

        conn_frame = tk.Frame(topbar, bg=COLORS["bg_dark"])
        conn_frame.pack(side="right", padx=16)
        self.conn_indicator = tk.Label(
            conn_frame,
            text="●",
            font=("Segoe UI", 12),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"],
        )
        self.conn_indicator.pack(side="left", padx=(0, 6))
        tk.Label(
            conn_frame,
            text=f"{ORACLE_USUARIO}@{ORACLE_HOST}",
            font=("Segoe UI", 9),
            fg=COLORS["text_secondary"],
            bg=COLORS["bg_dark"],
        ).pack(side="left")

        tk.Frame(self.root, bg=COLORS["border"], height=1).pack(fill="x")

        self.paned = tk.PanedWindow(
            self.root,
            orient="vertical",
            bg=COLORS["bg_dark"],
            sashwidth=6,
            sashrelief="flat",
            borderwidth=0,
        )
        self.paned.pack(fill="both", expand=True)

        editor_panel = tk.Frame(self.paned, bg=COLORS["bg_dark"])
        self.paned.add(editor_panel, minsize=150, height=260)

        toolbar = tk.Frame(editor_panel, bg=COLORS["bg_medium"], height=44)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        toolbar_inner = tk.Frame(toolbar, bg=COLORS["bg_medium"])
        toolbar_inner.pack(fill="x", padx=10, pady=6)

        self.btn_execute = ttk.Button(
            toolbar_inner,
            text="▶  Executar  (F5)",
            style="Green.TButton",
            command=self._execute_query,
        )
        self.btn_execute.pack(side="left", padx=(0, 6))

        self.btn_stop = ttk.Button(
            toolbar_inner,
            text="■  Parar",
            style="Red.TButton",
            command=self._stop_query,
            state="disabled",
        )
        self.btn_stop.pack(side="left", padx=(0, 16))

        tk.Frame(toolbar_inner, bg=COLORS["border"], width=1, height=28).pack(
            side="left", padx=8, fill="y"
        )

        ttk.Button(
            toolbar_inner,
            text="🗑  Limpar",
            style="Flat.TButton",
            command=self._clear_editor,
        ).pack(side="left", padx=(0, 4))

        ttk.Button(
            toolbar_inner,
            text="📋  Histórico",
            style="Flat.TButton",
            command=self._show_history,
        ).pack(side="left", padx=(0, 4))

        ttk.Button(
            toolbar_inner,
            text="📝  Templates",
            style="Flat.TButton",
            command=self._show_templates,
        ).pack(side="left", padx=(0, 4))

        tk.Label(
            toolbar_inner,
            text="Ctrl+Enter: Executar  |  Ctrl+E: Exportar Excel  |  Ctrl+L: Limpar",
            font=("Segoe UI", 8),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_medium"],
        ).pack(side="right")

        editor_container = tk.Frame(editor_panel, bg=COLORS["bg_dark"])
        editor_container.pack(fill="both", expand=True, padx=8, pady=(4, 4))

        self.line_numbers = tk.Text(
            editor_container,
            width=4,
            padx=8,
            pady=8,
            bg=COLORS["bg_light"],
            fg=COLORS["text_muted"],
            font=("Consolas", 11),
            relief="flat",
            state="disabled",
            cursor="arrow",
            selectbackground=COLORS["bg_light"],
            selectforeground=COLORS["text_muted"],
            borderwidth=0,
            highlightthickness=0,
        )
        self.line_numbers.pack(side="left", fill="y")

        self.sql_editor = tk.Text(
            editor_container,
            wrap="none",
            padx=10,
            pady=8,
            bg=COLORS["bg_input"],
            fg=COLORS["text_primary"],
            insertbackground=COLORS["accent"],
            selectbackground=COLORS["selection"],
            selectforeground="#ffffff",
            font=("Consolas", 11),
            relief="flat",
            undo=True,
            autoseparators=True,
            borderwidth=0,
            highlightthickness=1,
            highlightcolor=COLORS["accent"],
            highlightbackground=COLORS["border"],
        )
        self.sql_editor.pack(side="left", fill="both", expand=True)

        editor_scroll = ttk.Scrollbar(
            editor_container,
            orient="vertical",
            command=self.sql_editor.yview,
            style="Custom.Vertical.TScrollbar",
        )
        editor_scroll.pack(side="right", fill="y")
        self.sql_editor.configure(yscrollcommand=editor_scroll.set)

        self.sql_editor.bind("<KeyRelease>", self._on_editor_change)
        self.sql_editor.bind("<MouseWheel>", self._on_editor_change)
        self.sql_editor.bind("<Tab>", self._on_tab)

        self.sql_editor.insert(
            "1.0", "SELECT * FROM PCFILIAL ORDER BY CODIGO FETCH FIRST 50 ROWS ONLY"
        )
        self._on_editor_change()

        results_panel = tk.Frame(self.paned, bg=COLORS["bg_dark"])
        self.paned.add(results_panel, minsize=200)

        results_toolbar = tk.Frame(results_panel, bg=COLORS["bg_medium"], height=42)
        results_toolbar.pack(fill="x")
        results_toolbar.pack_propagate(False)

        results_toolbar_inner = tk.Frame(results_toolbar, bg=COLORS["bg_medium"])
        results_toolbar_inner.pack(fill="x", padx=10, pady=6)

        tk.Label(
            results_toolbar_inner,
            text="📊 Resultados",
            font=("Segoe UI", 11, "bold"),
            fg=COLORS["text_primary"],
            bg=COLORS["bg_medium"],
        ).pack(side="left")

        self.results_info = tk.Label(
            results_toolbar_inner,
            text="",
            font=("Segoe UI", 9),
            fg=COLORS["text_secondary"],
            bg=COLORS["bg_medium"],
        )
        self.results_info.pack(side="left", padx=(16, 0))

        actions_frame = tk.Frame(results_toolbar_inner, bg=COLORS["bg_medium"])
        actions_frame.pack(side="right")

        tk.Label(
            actions_frame,
            text="🔍",
            font=("Segoe UI", 10),
            bg=COLORS["bg_medium"],
            fg=COLORS["text_secondary"],
        ).pack(side="left")
        self.filter_entry = tk.Entry(
            actions_frame,
            textvariable=self.filter_text,
            width=25,
            font=("Segoe UI", 9),
            bg=COLORS["bg_input"],
            fg=COLORS["text_primary"],
            insertbackground=COLORS["accent"],
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightcolor=COLORS["accent"],
            highlightbackground=COLORS["border"],
        )
        self.filter_entry.pack(side="left", padx=(4, 12))

        tk.Frame(actions_frame, bg=COLORS["border"], width=1, height=24).pack(
            side="left", padx=6, fill="y"
        )

        ttk.Button(
            actions_frame,
            text="📥 Excel",
            style="Accent.TButton",
            command=self._export_excel,
        ).pack(side="left", padx=(6, 4))
        ttk.Button(
            actions_frame, text="📄 CSV", style="Flat.TButton", command=self._export_csv
        ).pack(side="left", padx=(0, 4))
        ttk.Button(
            actions_frame,
            text="📋 Copiar",
            style="Flat.TButton",
            command=self._copy_all,
        ).pack(side="left")

        results_container = tk.Frame(results_panel, bg=COLORS["bg_dark"])
        results_container.pack(fill="both", expand=True, padx=8, pady=(4, 4))

        self.tree = ttk.Treeview(
            results_container,
            style="Custom.Treeview",
            show="headings",
            selectmode="extended",
        )
        self.tree.pack(side="left", fill="both", expand=True)

        tree_vscroll = ttk.Scrollbar(
            results_container,
            orient="vertical",
            command=self.tree.yview,
            style="Custom.Vertical.TScrollbar",
        )
        tree_vscroll.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=tree_vscroll.set)

        tree_hscroll = ttk.Scrollbar(
            results_panel,
            orient="horizontal",
            command=self.tree.xview,
            style="Custom.Horizontal.TScrollbar",
        )
        tree_hscroll.pack(fill="x", padx=8, pady=(0, 2))
        self.tree.configure(xscrollcommand=tree_hscroll.set)

        self.tree.tag_configure("odd", background=COLORS["row_odd"])
        self.tree.tag_configure("even", background=COLORS["row_even"])

        self.tree_menu = tk.Menu(
            self.root,
            tearoff=0,
            bg=COLORS["bg_light"],
            fg=COLORS["text_primary"],
            font=("Segoe UI", 10),
            activebackground=COLORS["accent"],
            activeforeground="#ffffff",
            borderwidth=0,
        )
        self.tree_menu.add_command(
            label="📋  Copiar Seleção", command=self._copy_selection
        )
        self.tree_menu.add_command(label="📋  Copiar Tudo", command=self._copy_all)
        self.tree_menu.add_separator()
        self.tree_menu.add_command(
            label="📥  Exportar Excel", command=self._export_excel
        )
        self.tree_menu.add_command(label="📄  Exportar CSV", command=self._export_csv)
        self.tree.bind("<Button-3>", self._show_tree_menu)

        status_bar = tk.Frame(self.root, bg=COLORS["bg_medium"], height=30)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        self.status_icon = tk.Label(
            status_bar,
            text="●",
            font=("Segoe UI", 10),
            fg=COLORS["accent_green"],
            bg=COLORS["bg_medium"],
        )
        self.status_icon.pack(side="left", padx=(12, 4))

        self.status_label = tk.Label(
            status_bar,
            text="Pronto",
            font=("Segoe UI", 9),
            fg=COLORS["text_secondary"],
            bg=COLORS["bg_medium"],
            anchor="w",
        )
        self.status_label.pack(side="left", fill="x", expand=True)

        self.row_count_label = tk.Label(
            status_bar,
            text="",
            font=("Segoe UI", 9),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_medium"],
        )
        self.row_count_label.pack(side="right", padx=12)

    def _on_editor_change(self, event=None):
        self._update_line_numbers()
        self._highlight_sql()

    def _update_line_numbers(self):
        self.line_numbers.configure(state="normal")
        self.line_numbers.delete("1.0", "end")
        line_count = int(self.sql_editor.index("end-1c").split(".")[0])
        line_nums = "\n".join(str(i) for i in range(1, line_count + 1))
        self.line_numbers.insert("1.0", line_nums)
        self.line_numbers.configure(state="disabled")
        self.line_numbers.yview_moveto(self.sql_editor.yview()[0])

    def _highlight_sql(self):
        self.sql_editor.tag_remove("keyword", "1.0", "end")
        self.sql_editor.tag_remove("string", "1.0", "end")
        self.sql_editor.tag_remove("number", "1.0", "end")
        self.sql_editor.tag_remove("comment", "1.0", "end")

        self.sql_editor.tag_configure(
            "keyword", foreground="#d35400", font=("Consolas", 11, "bold")
        )
        self.sql_editor.tag_configure("string", foreground="#27ae60")
        self.sql_editor.tag_configure("number", foreground="#e67e22")
        self.sql_editor.tag_configure(
            "comment", foreground="#95a5a6", font=("Consolas", 11, "italic")
        )

        content = self.sql_editor.get("1.0", "end")

        for kw in SQL_KEYWORDS:
            start = "1.0"
            while True:
                pos = self.sql_editor.search(
                    r"\m" + kw + r"\M", start, "end", nocase=True, regexp=True
                )
                if not pos:
                    break
                end_pos = f"{pos}+{len(kw)}c"
                self.sql_editor.tag_add("keyword", pos, end_pos)
                start = end_pos

        start = "1.0"
        in_string = False
        string_start = None
        while True:
            pos = self.sql_editor.search("'", start, "end")
            if not pos:
                break
            if not in_string:
                string_start = pos
                in_string = True
            else:
                end_pos = f"{pos}+1c"
                self.sql_editor.tag_add("string", string_start, end_pos)
                in_string = False
            start = f"{pos}+1c"

        lines = content.split("\n")
        for i, line in enumerate(lines, 1):
            idx = line.find("--")
            if idx >= 0:
                self.sql_editor.tag_add("comment", f"{i}.{idx}", f"{i}.end")

    def _on_tab(self, event):
        self.sql_editor.insert("insert", "    ")
        return "break"

    def _clear_editor(self):
        self.sql_editor.delete("1.0", "end")
        self._on_editor_change()

    def _execute_query(self):
        if self.is_running:
            return

        sql = self.sql_editor.get("1.0", "end").strip()
        if not sql:
            self._set_status("Nenhuma query para executar.", "warning")
            return

        if not oracle_available:
            messagebox.showerror(
                "Erro",
                "Módulo oracledb não está disponível.\nInstale com: pip install oracledb",
            )
            return

        self.is_running = True
        self.btn_execute.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._set_status("Executando query...", "running")
        self._clear_results()

        self._query_thread = threading.Thread(
            target=self._run_query, args=(sql,), daemon=True
        )
        self._query_thread.start()

    def _run_query(self, sql):
        start_time = time.time()
        try:
            conn = conectar()
            self.connection = conn
            cursor = conn.cursor()
            cursor.arraysize = 1000
            cursor.execute(sql)

            if cursor.description:
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                elapsed = time.time() - start_time

                display_rows = []
                for row in rows:
                    display_row = []
                    for val in row:
                        if val is None:
                            display_row.append("")
                        elif isinstance(val, datetime):
                            display_row.append(val.strftime("%d/%m/%Y %H:%M:%S"))
                        elif isinstance(val, (bytes, bytearray)):
                            display_row.append(f"<BLOB {len(val)} bytes>")
                        else:
                            try:
                                s = str(val)
                                display_row.append(s)
                            except Exception:
                                try:
                                    raw = val.__str__()
                                    if isinstance(raw, bytes):
                                        display_row.append(
                                            raw.decode("utf-8", errors="replace")
                                        )
                                    else:
                                        display_row.append(repr(val))
                                except Exception:
                                    display_row.append(repr(val))
                    display_rows.append(display_row)

                self.root.after(
                    0, self._display_results, columns, display_rows, elapsed, len(rows)
                )
                self._add_to_history(sql, len(rows), elapsed)
            else:
                affected = cursor.rowcount
                conn.commit()
                elapsed = time.time() - start_time
                self.root.after(
                    0,
                    self._set_status,
                    f"Comando executado com sucesso. {affected} linha(s) afetada(s). ({elapsed:.2f}s)",
                    "success",
                )

            cursor.close()

        except Exception as e:
            elapsed = time.time() - start_time
            try:
                error_msg = str(e)
            except Exception:
                try:
                    error_msg = e.args[0] if e.args else repr(e)
                    if isinstance(error_msg, bytes):
                        error_msg = error_msg.decode("utf-8", errors="replace")
                    else:
                        error_msg = repr(e)
                except Exception:
                    error_msg = repr(e)
            self.root.after(0, self._set_status, f"ERRO: {error_msg}", "error")
            self.root.after(
                0, lambda m=error_msg: messagebox.showerror("Erro na Query", m)
            )

        finally:
            if self.connection:
                try:
                    self.connection.close()
                except:
                    pass
                self.connection = None
            self.root.after(0, self._finish_query)

    def _finish_query(self):
        self.is_running = False
        self.btn_execute.configure(state="normal")
        self.btn_stop.configure(state="disabled")

    def _stop_query(self):
        if self.connection:
            try:
                self.connection.cancel()
            except:
                pass
        self._set_status("Query cancelada pelo usuário.", "warning")

    def _clear_results(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()
        self.current_columns = []
        self.current_data = []
        self.results_info.configure(text="")
        self.row_count_label.configure(text="")

    def _display_results(self, columns, rows, elapsed, total_rows):
        self.current_columns = columns
        self.current_data = rows

        self.tree["columns"] = columns
        for i, col in enumerate(columns):
            self.tree.heading(
                col, text=f"  {col}  ", command=lambda c=col: self._sort_by_column(c)
            )

            max_width = max(len(col) * 12 + 30, 120)
            sample = rows[:100] if len(rows) > 100 else rows
            for row in sample:
                if i < len(row):
                    cell_width = len(str(row[i])) * 9 + 30
                    max_width = max(max_width, min(cell_width, 500))
            self.tree.column(col, width=max_width, minwidth=100, anchor="w")

        self._populate_tree(rows)

        self.results_info.configure(
            text=f"   {total_rows} registro(s)  ·  {elapsed:.3f}s  ·  {len(columns)} coluna(s)"
        )
        self.row_count_label.configure(text=f"{total_rows} linhas")
        self._set_status(
            f"Query executada com sucesso! {total_rows} registro(s) em {elapsed:.3f}s",
            "success",
        )

    def _populate_tree(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", values=row, tags=(tag,))

    def _sort_by_column(self, col):
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        col_idx = self.current_columns.index(col)

        def sort_key(row):
            val = row[col_idx] if col_idx < len(row) else ""
            try:
                return (0, float(val.replace(",", ".")))
            except (ValueError, AttributeError):
                return (1, str(val).lower())

        sorted_data = sorted(self.current_data, key=sort_key, reverse=self.sort_reverse)
        self._populate_tree(sorted_data)

        for c in self.current_columns:
            arrow = ""
            if c == col:
                arrow = " ▼" if self.sort_reverse else " ▲"
            self.tree.heading(c, text=c + arrow)

    def _on_filter_change(self, *args):
        if not self.current_data:
            return

        text = self.filter_text.get().lower().strip()
        if not text:
            self._populate_tree(self.current_data)
            self.row_count_label.configure(text=f"{len(self.current_data)} linhas")
            return

        filtered = []
        for row in self.current_data:
            for val in row:
                if text in str(val).lower():
                    filtered.append(row)
                    break

        self._populate_tree(filtered)
        self.row_count_label.configure(
            text=f"{len(filtered)} de {len(self.current_data)} linhas"
        )

    def _export_excel(self):
        if not self.current_data:
            messagebox.showinfo("Aviso", "Nenhum dado para exportar.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Erro",
                "Módulo openpyxl não instalado.\nInstale com: pip install openpyxl",
            )
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"query_resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title="Exportar para Excel",
        )

        if not filepath:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultado SQL"

            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="E67E22", end_color="E67E22", fill_type="solid"
            )
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            for col_idx, col_name in enumerate(self.current_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            data_font = Font(name="Calibri", size=10)
            even_fill = PatternFill(
                start_color="FEF3E2", end_color="FEF3E2", fill_type="solid"
            )
            odd_fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

            for row_idx, row_data in enumerate(self.current_data, 2):
                fill = even_fill if row_idx % 2 == 0 else odd_fill
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.fill = fill
                    cell.border = thin_border
                    try:
                        cell.value = float(value.replace(",", "."))
                        cell.number_format = "#,##0.00"
                    except (ValueError, AttributeError):
                        pass

            for col_idx, col_name in enumerate(self.current_columns, 1):
                max_len = len(col_name)
                for row_data in self.current_data[:200]:
                    if col_idx - 1 < len(row_data):
                        max_len = max(max_len, len(str(row_data[col_idx - 1])))
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = min(max_len + 4, 50)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            wb.save(filepath)
            self._set_status(f"Exportado para Excel: {filepath}", "success")
            messagebox.showinfo(
                "Sucesso",
                f"Arquivo exportado com sucesso!\n\n{filepath}\n\n{len(self.current_data)} registro(s)",
            )

        except Exception as e:
            messagebox.showerror("Erro ao Exportar", str(e))

    def _export_csv(self):
        if not self.current_data:
            messagebox.showinfo("Aviso", "Nenhum dado para exportar.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"query_resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            title="Exportar para CSV",
        )

        if not filepath:
            return

        try:
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(self.current_columns)
                writer.writerows(self.current_data)

            self._set_status(f"Exportado para CSV: {filepath}", "success")
            messagebox.showinfo(
                "Sucesso",
                f"CSV exportado!\n\n{filepath}\n\n{len(self.current_data)} registro(s)",
            )

        except Exception as e:
            messagebox.showerror("Erro ao Exportar", str(e))

    def _copy_selection(self):
        selected = self.tree.selection()
        if not selected:
            return

        lines = []
        lines.append("\t".join(self.current_columns))
        for item in selected:
            values = self.tree.item(item, "values")
            lines.append("\t".join(str(v) for v in values))

        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lines))
        self._set_status(
            f"{len(selected)} linha(s) copiada(s) para a área de transferência.",
            "success",
        )

    def _copy_all(self):
        if not self.current_data:
            return

        lines = []
        lines.append("\t".join(self.current_columns))
        for row in self.current_data:
            lines.append("\t".join(str(v) for v in row))

        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lines))
        self._set_status(
            f"{len(self.current_data)} linha(s) copiada(s) para a área de transferência.",
            "success",
        )

    def _show_tree_menu(self, event):
        try:
            self.tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.tree_menu.grab_release()

    def _load_history(self):
        try:
            if os.path.exists(HISTORY_FILE):
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    self.query_history = json.load(f)
        except:
            self.query_history = []

    def _save_history(self):
        try:
            with open(HISTORY_FILE, "w", encoding="utf-8") as f:
                json.dump(self.query_history[-50:], f, ensure_ascii=False, indent=2)
        except:
            pass

    def _add_to_history(self, sql, rows, elapsed):
        entry = {
            "sql": sql,
            "rows": rows,
            "elapsed": round(elapsed, 3),
            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        }
        self.query_history.append(entry)
        self._save_history()

    def _show_history(self):
        if not self.query_history:
            messagebox.showinfo("Histórico", "Nenhuma query no histórico.")
            return

        hist_win = tk.Toplevel(self.root)
        hist_win.title("Histórico de Queries")
        hist_win.geometry("800x500")
        hist_win.configure(bg=COLORS["bg_dark"])
        hist_win.transient(self.root)
        hist_win.grab_set()

        tk.Label(
            hist_win,
            text="📋  Histórico de Queries",
            font=("Segoe UI", 14, "bold"),
            fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"],
        ).pack(pady=(16, 8))
        tk.Label(
            hist_win,
            text="Clique duas vezes para carregar a query no editor",
            font=("Segoe UI", 9),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"],
        ).pack(pady=(0, 8))

        container = tk.Frame(hist_win, bg=COLORS["bg_dark"])
        container.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        listbox = tk.Listbox(
            container,
            bg=COLORS["bg_input"],
            fg=COLORS["text_primary"],
            font=("Consolas", 10),
            selectbackground=COLORS["accent"],
            selectforeground="#ffffff",
            borderwidth=0,
            highlightthickness=1,
            highlightcolor=COLORS["accent"],
            highlightbackground=COLORS["border"],
        )
        listbox.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(container, orient="vertical", command=listbox.yview)
        sb.pack(side="right", fill="y")
        listbox.configure(yscrollcommand=sb.set)

        for entry in reversed(self.query_history):
            sql_preview = entry["sql"][:100].replace("\n", " ")
            listbox.insert(
                "end",
                f"[{entry['timestamp']}]  {entry['rows']} rows  ({entry['elapsed']}s)  │  {sql_preview}",
            )

        def on_select(event):
            idx = listbox.curselection()
            if idx:
                real_idx = len(self.query_history) - 1 - idx[0]
                sql = self.query_history[real_idx]["sql"]
                self.sql_editor.delete("1.0", "end")
                self.sql_editor.insert("1.0", sql)
                self._on_editor_change()
                hist_win.destroy()

        listbox.bind("<Double-Button-1>", on_select)

        ttk.Button(
            hist_win, text="Fechar", style="Flat.TButton", command=hist_win.destroy
        ).pack(pady=(0, 16))

    def _show_templates(self):
        templates = [
            (
                "Todas as Filiais",
                "SELECT CODIGO, RAZAOSOCIAL, CGCFILIAL AS CNPJ, CIDADE, UF\nFROM PCFILIAL\nORDER BY CODIGO",
            ),
            (
                "Produtos (Top 100)",
                "SELECT CODPROD, DESCRICAO, EMBALAGEM, UNIDADE, CODAUXILIAR\nFROM PCPRODUT\nORDER BY CODPROD\nFETCH FIRST 100 ROWS ONLY",
            ),
            (
                "Fornecedores (Top 100)",
                "SELECT CODFORNEC, FORNECEDOR, CGC, CIDADE, UF\nFROM PCFORNEC\nORDER BY CODFORNEC\nFETCH FIRST 100 ROWS ONLY",
            ),
            (
                "Clientes (Top 100)",
                "SELECT CODCLI, CLIENTE, CGCENT, CIDADE, ESTENT\nFROM PCCLIENT\nORDER BY CODCLI\nFETCH FIRST 100 ROWS ONLY",
            ),
            (
                "Tabelas do Schema",
                "SELECT TABLE_NAME, NUM_ROWS, LAST_ANALYZED\nFROM USER_TABLES\nORDER BY TABLE_NAME",
            ),
            (
                "Colunas de uma Tabela",
                "SELECT COLUMN_NAME, DATA_TYPE, DATA_LENGTH, NULLABLE\nFROM USER_TAB_COLUMNS\nWHERE TABLE_NAME = 'PCPRODUT'\nORDER BY COLUMN_ID",
            ),
            (
                "Estoque por Filial",
                "SELECT E.CODPROD, P.DESCRICAO, E.CODFILIAL, E.QTESTGER, E.QTRESERV\nFROM PCEST E\nJOIN PCPRODUT P ON P.CODPROD = E.CODPROD\nWHERE E.CODFILIAL = 1\nORDER BY P.DESCRICAO\nFETCH FIRST 100 ROWS ONLY",
            ),
            (
                "Vendas do Dia",
                "SELECT NUMPED, CODCLI, VLTOTAL, DATA, CODFILIAL\nFROM PCPEDC\nWHERE TRUNC(DATA) = TRUNC(SYSDATE)\nORDER BY NUMPED DESC\nFETCH FIRST 100 ROWS ONLY",
            ),
        ]

        tmpl_win = tk.Toplevel(self.root)
        tmpl_win.title("Templates SQL")
        tmpl_win.geometry("700x480")
        tmpl_win.configure(bg=COLORS["bg_dark"])
        tmpl_win.transient(self.root)
        tmpl_win.grab_set()

        tk.Label(
            tmpl_win,
            text="📝  Templates SQL — Winthor",
            font=("Segoe UI", 14, "bold"),
            fg=COLORS["text_primary"],
            bg=COLORS["bg_dark"],
        ).pack(pady=(16, 4))
        tk.Label(
            tmpl_win,
            text="Clique em um template para carregar no editor",
            font=("Segoe UI", 9),
            fg=COLORS["text_muted"],
            bg=COLORS["bg_dark"],
        ).pack(pady=(0, 12))

        canvas = tk.Canvas(tmpl_win, bg=COLORS["bg_dark"], highlightthickness=0)
        scrollbar = ttk.Scrollbar(tmpl_win, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=COLORS["bg_dark"])

        scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=16)
        scrollbar.pack(side="right", fill="y")

        for name, sql in templates:
            frame = tk.Frame(
                scroll_frame, bg=COLORS["bg_light"], cursor="hand2", padx=14, pady=10
            )
            frame.pack(fill="x", pady=4, padx=4)

            lbl_name = tk.Label(
                frame,
                text=name,
                font=("Segoe UI", 11, "bold"),
                fg=COLORS["text_primary"],
                bg=COLORS["bg_light"],
                anchor="w",
                cursor="hand2",
            )
            lbl_name.pack(fill="x")

            lbl_sql = tk.Label(
                frame,
                text=sql,
                font=("Consolas", 9),
                fg=COLORS["text_secondary"],
                bg=COLORS["bg_light"],
                anchor="w",
                justify="left",
                cursor="hand2",
            )
            lbl_sql.pack(fill="x", pady=(4, 0))

            def load_template(s=sql, w=tmpl_win):
                self.sql_editor.delete("1.0", "end")
                self.sql_editor.insert("1.0", s)
                self._on_editor_change()
                w.destroy()

            for widget in (frame, lbl_name, lbl_sql):
                widget.bind("<Button-1>", lambda e, s=sql: load_template(s))
                widget.bind(
                    "<Enter>",
                    lambda e, f=frame, ln=lbl_name, ls=lbl_sql: (
                        f.configure(bg=COLORS["accent"]),
                        ln.configure(bg=COLORS["accent"]),
                        ls.configure(bg=COLORS["accent"]),
                    ),
                )
                widget.bind(
                    "<Leave>",
                    lambda e, f=frame, ln=lbl_name, ls=lbl_sql: (
                        f.configure(bg=COLORS["bg_light"]),
                        ln.configure(bg=COLORS["bg_light"]),
                        ls.configure(bg=COLORS["bg_light"]),
                    ),
                )

    def _set_status(self, msg, level="info"):
        colors = {
            "info": COLORS["text_secondary"],
            "success": COLORS["accent_green"],
            "warning": COLORS["accent_orange"],
            "error": COLORS["accent_red"],
            "running": COLORS["accent"],
        }
        icons = {
            "info": "●",
            "success": "✓",
            "warning": "⚠",
            "error": "✕",
            "running": "◌",
        }
        self.status_label.configure(
            text=msg, fg=colors.get(level, COLORS["text_secondary"])
        )
        self.status_icon.configure(
            text=icons.get(level, "●"), fg=colors.get(level, COLORS["text_secondary"])
        )

        if level == "success":
            self.conn_indicator.configure(fg=COLORS["accent_green"])
        elif level == "error":
            self.conn_indicator.configure(fg=COLORS["accent_red"])


if __name__ == "__main__":
    root = tk.Tk()

    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass

    app = ViewOracleApp(root)
    root.mainloop()
